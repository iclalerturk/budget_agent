from budget_llm_project.src.services.excel_reader import ExcelReader
from budget_llm_project.src.services.document_processor import DocumentProcessor
from budget_llm_project.src.services.index_manager import IndexManager
from budget_llm_project.src.llm.llm_client import LLMClient
from budget_llm_project.src.utils.query_engine import QueryEngine
from openpyxl import load_workbook


if __name__ == "__main__":
    # 1) Client
    client = LLMClient(
        api_key="****************",
        endpoint="******************",
    )

    # 2) Excel oku
    reader = ExcelReader(".venv\\budget_llm_project\\data\\veriler.xlsx")
    documents_raw = reader.read_excel_with_metadata()

    # 3) Doc list oluştur
    processor = DocumentProcessor()
    doc_list = processor.create_documents_with_headers(documents_raw, header_row=1)

    # 4) Index Manager
    index_manager = IndexManager()
    index_data, doc_list_data, embeddings_data = index_manager.build_index_for_sheet(doc_list, "Data")

    # 5) Query Engine
    engine = QueryEngine(client)

    print("Excel Sheet Names:", [ws for ws in load_workbook(reader.file_path).sheetnames])
    print("Metadata Sheets:", list(set(doc["metadata"]["sheet"] for doc in doc_list)))

    while True:
        soru = input("Lütfen sorunuzu girin: ")
        print(engine.smart_query(soru, doc_list_data, index_data))
