from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class ExcelReader:
    def __init__(self, file_path):
        self.file_path = file_path

    def read_excel_with_metadata(self):
        wb = load_workbook(self.file_path, data_only=True)
        documents = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            merged_map = {}
            for merged_range in ws.merged_cells.ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds
                value = ws.cell(row=min_row, column=min_col).value
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        merged_map[(r, c)] = value
            for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
                row_data = []
                for col_idx, cell in enumerate(row, start=1):
                    value = merged_map.get((row_idx, col_idx), cell.value)
                    if value is None:
                        display_val = "EMPTY"
                    else:
                        display_val = value
                    row_data.append({
                        "value": display_val,
                        "type": str(type(display_val)),
                        "sheet": sheet,
                        "row": row_idx,
                        "col": col_idx,
                        "address": f"{get_column_letter(col_idx)}{row_idx}"
                    })
                documents.append(row_data)
        return documents
